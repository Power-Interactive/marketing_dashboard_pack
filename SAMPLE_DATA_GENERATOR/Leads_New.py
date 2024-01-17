#   python3 dashLeadMake.py \
#       specSheet='サンプルデータ作成仕様書_20230629.xlsx' \
#       sampleLead='sample_ms-Leads.xlsx' \
#       leadData='leadData.csv' \
#       newLead='leadDataNew.csv' \
#       personData='Dummy_Data_50000.xlsx' \
#       startdate=2021-06-30T15:21:43Z
#       [merge]


import os
import sys
import re
import openpyxl
import csv
import random
import datetime as dt
import zoneinfo
import numpy as np
import json
import string
import pprint
import pathlib
import pidbcommon


args                = sys.argv
flg_debug           = False
dict_param          = {}
baseDateCreated     = ''
dt_baseDateUpdated  = ''
dt_baseDateCreated  = ''
dt_baseDate         = ''
dict_configData     = {}
int_Number          = 30
dict_param['merge'] = True
int_eventIndex      = 0
dict_eventSource    = {}

for arg in args:
    if arg == 'Leads_New.py':
        continue
    elif re.match(r'confi?g?u?r?e?=', arg, re.IGNORECASE):
        elements = re.split('=', arg)
        dict_param['configuration'] = elements[1]
    elif re.match(r'file_New=', arg, re.IGNORECASE):
        elements = re.split('=', arg)
        dict_param['file_New'] = elements[1]
    else:
        print("Unknown argument : "+arg, file=sys.stderr)
        sys.exit()

if 'configuration' in dict_param:
    #   保存情報の取得
    with open(dict_param['configuration']) as f:
        dict_configData = json.load(f)
else:
    print("'conf(iguration)=' parameter not specified.", file=sys.stderr)
    sys.exit()

if not 'actfileprefix' in dict_configData['common']:
    print("saveData_Lead error : 'actfileprefix' key in JSON fike not found", file=sys.stderr)
    sys.exit()

if 'referenceSheet' in dict_configData['common']:
    print(dict_configData['common']['referenceSheet'],"OK")
else:
    print("'referenceSheet' key in JSON file not found", file=sys.stderr)
    sys.exit()

if 'number' in dict_configData['Leads_New']:
    int_Number = dict_configData['Leads_New']['number']
else:
    print("[information] Number of generate data : 'number' key in JSON file not found. 'number' was set to 30.")

if 'file_New' in dict_param:
    pass
elif 'file_New' in dict_configData['Leads_New']:
    dict_param['file_New'] = dict_configData["Leads_New"]['file_New']
else:
    print("'file_New' parameter is not found.", file=sys.stderr)
    sys.exit()

#   'CreatedAt'の処理
if 'CreatedAt' in dict_configData['Leads_New']:
    print(dict_configData['Leads_New']['CreatedAt'],"OK")
    try:
        dt_tmp = dt.datetime.fromisoformat(dict_configData['Leads_New']['CreatedAt'])
    except ValueError:
        print("Date format error : 'CreatedAt' -> "+dict_configData['Leads_New']['CreatedAt'], file=sys.stderr)
        sys.exit()
    str_CreatedAt = dict_configData['Leads_New']['CreatedAt'].replace('Z', '+00:00')
    dt_baseDate = dt.datetime.fromisoformat(str_CreatedAt)
    td = dt.timedelta(days=1)
    dt_baseDate += td
    str_baseDay = dt_baseDate.date().isoformat()
    re.sub('T.*', '', str_baseDay)
else:
    print("saveData_Lead error : 'CreatedAt' key in JSON file not found", file=sys.stderr)
    sys.exit()

if 'event_month' in dict_configData['Leads_New']:
    for idx, int_eventMonth in enumerate(dict_configData['Leads_New']['event_month']):
        print(">>>> int_eventMonth :", int_eventMonth, "   >>>> dt_baseDate.month :", dt_baseDate.month)
        if int_eventMonth == dt_baseDate.month:
            int_Number = int(int_Number * 1.5)
            #   イベント/展示が始まるNoを求める。リードソースマスターの並びに左右される。今回は+11
            int_eventIndex = idx + 11
            break



wb = openpyxl.load_workbook(dict_configData['common']['referenceSheet'])

#   リードソースマスターを辞書のリストに格納
ws = wb['リードソースマスター']
header_cells = ws[1]
list_LeadSourceMaster = []
for row in ws.iter_rows(min_row=2, max_row=31):
    row_dic = {}
    # セルの値を「key-value」で登録
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_LeadSourceMaster.append(row_dic)
print('-------------------------------- リードソースマスタ')
pprint.pprint(list_LeadSourceMaster)
print(">>>> int_eventIndex :", int_eventIndex)
if int_eventIndex != 0:
    for row in list_LeadSourceMaster:
        if int(row['№']) == int_eventIndex:
            dict_eventSource = row
            break
pprint.pprint(dict_eventSource)

#   アカウントマスターを辞書のリストに格納
ws = wb['sample_sfa.Account']
header_cells = ws[1]
list_AccountMaster = []
dict_AccountMaster = {'J1':[], 'J2':[], 'J3':[]}
for ii, row in enumerate(ws.iter_rows(min_row=2, max_row=61)):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
        list_AccountMaster.append(row_dic)
    if row_dic['Rank'] == 'A':
        dict_AccountMaster['J1'].append(row_dic)
    elif row_dic['Rank'] == 'B':
        dict_AccountMaster['J2'].append(row_dic)
    else:
        dict_AccountMaster['J3'].append(row_dic)

#   部署マスター（Department）を辞書のリストに格納
ws = wb['部署マスター']
header_cells = ws[1]
list_JLeagueDepartmentMaster = []
for row in ws.iter_rows(min_row=2, max_row=71):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_JLeagueDepartmentMaster.append(row_dic)

#   事業所マスター（Office）を辞書のリストに格納
ws = wb['事業所マスター']
header_cells = ws[1]
list_JLeagueOfficeMaster = []
for row in ws.iter_rows(min_row=2, max_row=11):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_JLeagueOfficeMaster.append(row_dic)

#   役職マスターを辞書のリストに格納
ws = wb['役職マスター']
header_cells = ws[1]
list_PositionMaster = []
for row in ws.iter_rows(min_row=2, max_row=18):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_PositionMaster.append(row_dic)

#   行動スコアマスターを辞書のリストに格納
ws = wb['行動スコアマスター']
header_cells = ws[1]
list_ActivityScoreMaster = []
for row in ws.iter_rows(min_row=2, max_row=6):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_ActivityScoreMaster.append(row_dic)

#   Contactを辞書のリストに格納
ws = wb['Contact']
header_cells = ws[1]
list_Contact = []
for row in ws.iter_rows(min_row=2, max_row=2104, min_col=1, max_col=1):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_Contact.append(row_dic)

ws = wb['ClickEmail']
header_cells = ws[1]
list_ClickEmailSample = []
for row in ws.iter_rows(min_row=2, max_row=59):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_ClickEmailSample.append(row_dic)

ws = wb['ClickLink']
header_cells = ws[1]
list_ClickLinkSample = []
for row in ws.iter_rows(min_row=2, max_row=355):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_ClickLinkSample.append(row_dic)

ws = wb['FillOutForm']
header_cells = ws[1]
list_FillOutFormSample = []
for row in ws.iter_rows(min_row=2, max_row=3717):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_FillOutFormSample.append(row_dic)

ws = wb['OpenEmail']
header_cells = ws[1]
list_OpenEmailSample = []
for row in ws.iter_rows(min_row=2, max_row=82):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_OpenEmailSample.append(row_dic)

ws = wb['VisitWebPage']
header_cells = ws[1]
list_VisitWebpageSample = []
for row in ws.iter_rows(min_row=2, max_row=497):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_VisitWebpageSample.append(row_dic)

#   氏名メールデータを辞書のリストに格納
list_PersonData = []
wb = openpyxl.load_workbook(dict_configData["Leads_New"]['persondata'])
max_row_PersonData = 10001
ws = wb['dummy']
header_cells = ws[1]
for row in ws.iter_rows(min_row=2, max_row=10001):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_PersonData.append(row_dic)
ws = wb['dummy (1)']
header_cells = ws[1]
for row in ws.iter_rows(min_row=2, max_row=10001):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_PersonData.append(row_dic)
ws = wb['dummy (2)']
header_cells = ws[1]
for row in ws.iter_rows(min_row=2, max_row=10001):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_PersonData.append(row_dic)
ws = wb['dummy (3)']
header_cells = ws[1]
for row in ws.iter_rows(min_row=2, max_row=10001):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_PersonData.append(row_dic)
ws = wb['dummy (4)']
header_cells = ws[1]
for row in ws.iter_rows(min_row=2, max_row=10001):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_PersonData.append(row_dic)

list_VisitWebpage       = []
list_FillOutForm        = []
list_ClickLink          = []
list_OpenEmail          = []
list_ClickEmail         = []
list_LeadDataNew        = []
list_ChangeDataValue    = []
repeatCount = random.randint(0, int_Number)
if dict_configData['Leads_New']['period'] == 1:
    list_dd = [0] * repeatCount
else:
    list_dd = [int(n/100) for n in pidbcommon.rand_ints_nodup(0, dict_configData['Leads_New']['period'] * 100, repeatCount)]
list_dd.sort()

for ii in range(repeatCount):
    dict_LeadData   = {'customer_id': '', 'CreatedAt': '', 'LeadSource1': '', 'LeadSource2': '', 'LeadSource3': ' ', 'lifecycle_status': '', 'Company': '', 'Department': '', 'Title': '', 'lastname': '', 'firstname': '', 'sfa_account_id': '', 'sfa_contact_id':'', 'sfa_lead_id': '', 'sfa_type': '', 'score': '', 'BehaviorScore': '', 'DemographicScore': '', 'UpdatedAt': '', 'number_of_employee': ''}
    #   インクリメントなどによるセット
    dict_configData['Leads_New']['customer_id']     += 1
    dict_LeadData['customer_id']        = dict_configData['Leads_New']['customer_id']
    hr = '{:0=2}:'.format(random.randint(7, 22))
    mm = '{:0=2}:'.format(random.randint(0, 59))
    ss = '{:0=2}.'.format(random.randint(0, 59))
    ms = '{:0=6}'.format(random.randint(0, 999999))
    td = dt.timedelta(days=list_dd[ii])
    dt_CreatedAt                        = dt_baseDate + td
    str_CreatedAtJST                    = dt_CreatedAt.date().isoformat() + 'T' + hr + mm + ss + ms
    dt_CreatedAt                        = dt.datetime.fromisoformat(str_CreatedAtJST)
    dt_CreatedAtJST                     = dt_CreatedAt.replace(tzinfo=zoneinfo.ZoneInfo(key='Asia/Tokyo'))
    dt_CreatedAtUTC                     = dt_CreatedAtJST.astimezone(dt.timezone.utc)
    dict_LeadData['CreatedAt']          = dt_CreatedAtUTC.isoformat()
    dict_configData['Leads_New']['CreatedAt']        = dt_CreatedAtJST.isoformat()

    #   リードソースマスターよりセットする
    idx                                     = random.randint(0, len(list_LeadSourceMaster) - 1)
    if int_eventIndex == 0:
        dict_LeadData['LeadSource1']        = list_LeadSourceMaster[idx]['OpprtunitySource']
        dict_LeadData['LeadSource3']        = list_LeadSourceMaster[idx]['Detailed_Opp_Source']
    else:
        int_Mode = random.randint(0, 2)
        if int_Mode == 2:
            dict_LeadData['LeadSource1']    = dict_eventSource['OpprtunitySource']
            dict_LeadData['LeadSource3']    = dict_eventSource['Detailed_Opp_Source']
        else:
            dict_LeadData['LeadSource1']    = list_LeadSourceMaster[idx]['OpprtunitySource']
            dict_LeadData['LeadSource3']    = list_LeadSourceMaster[idx]['Detailed_Opp_Source']
    
    #   アカウントマスターよりセットする
    Jx                                      = np.random.choice(['J1', 'J2', 'J3'], p=[500/750, 200/750, 50/750])
    idx                                     = random.randint(0, len(dict_AccountMaster[Jx]) - 1)
    dict_LeadData['Company']                = dict_AccountMaster[Jx][idx]['Name']
    dict_LeadData['sfa_account_id']         = dict_AccountMaster[Jx][idx]['ID']
    dict_LeadData['number_of_employee']     = dict_AccountMaster[Jx][idx]['number_of_employee']
    
    #   部署マスターよりセットする
    idx                                     = random.randint(0, len(list_JLeagueOfficeMaster) - 1)
    dict_LeadData['Department']             = list_JLeagueDepartmentMaster[idx]['Department']
    
    #   役職マスターよりセットする
    idx                                     = random.randint(0, len(list_PositionMaster) - 1)
    dict_LeadData['Title']                  = list_PositionMaster[idx]['役職']
    dict_LeadData['DemographicScore']       = int(list_PositionMaster[idx]['属性スコア'])
    
    #   氏名メールマスターよりセットする
    dict_configData['Leads_New']['persondataCNT']  += 1
    if dict_configData['Leads_New']['persondataCNT'] > 49999:
        dict_configData['Leads_New']['persondataCNT'] = 1
    try:
        dict_LeadData['firstname']          = list_PersonData[dict_configData['Leads_New']['persondataCNT']]['姓']
    except IndexError:
        print("================= IndexError : {:06}".format(dict_configData['Leads_New']['persondataCNT']), file=sys.stderr)
        sys.exit()
    dict_LeadData['lastname']               = list_PersonData[dict_configData['Leads_New']['persondataCNT']]['名']

    #   行動スコアマスターマスターよりセットする
    idx_act                                 = random.randint(0, len(list_ActivityScoreMaster) - 1)
    dict_LeadData['BehaviorScore']          = int(list_ActivityScoreMaster[idx_act]['行動スコア（加点）'])

    #   Contactよりセットする
    idx                                     = random.randint(0, len(list_Contact) - 1)
    dict_LeadData['sfa_contact_id']         = list_Contact[idx]['id']

    #   計算によりセット
    dict_LeadData['score']                  = dict_LeadData['BehaviorScore'] + dict_LeadData['DemographicScore']

    #   固定セット
    dict_LeadData['lifecycle_status']       = "MCL"
    dict_LeadData['sfa_type']               = 'Contact'
    list_LeadDataNew.append(dict_LeadData)

    #   ChangeDataValueのセット
    dict_tmp                        = {}
    dict_configData['common']['ActivityId']  += 1
    dict_tmp['ActivityId']          = dict_configData['common']['ActivityId']
    dict_tmp['ActivityDate']        = dict_LeadData['CreatedAt']
    dict_tmp['AttributeName']       = 1292
    dict_tmp['AttributeNameValue']  = 'Lifecycle Status'
    dict_tmp['CampaignId']          = '5218'
    dict_tmp['LeadId']              = dict_LeadData['customer_id']
    dict_tmp['NewValue']            = dict_LeadData['lifecycle_status']
    dict_tmp['OldValue']            = 'ANONYMOUS'
    dict_tmp['Reason']              = ''
    list_ChangeDataValue.append(dict_tmp)

    #   Activityログのセット
    dict_tmp                        = {}
    dict_configData['common']['ActivityId']  += 1
    dict_tmp['ActivityId']          = dict_configData['common']['ActivityId']
    dict_tmp['ActivityDate']        = dict_LeadData['CreatedAt']
    dict_tmp['LeadId']              = dict_LeadData['customer_id']
    if list_ActivityScoreMaster[idx_act]['Activity Type'] == 'Visit Webpage':
        ii2                         = random.randint(0, len(list_VisitWebpageSample) - 1)
        dict_tmp['WebpageIDValue']  = list_VisitWebpageSample[ii2]['WebpageIDValue']
        dict_tmp['WebpageURL']      = list_VisitWebpageSample[ii2]['WebpageURL']
        list_VisitWebpage.append(dict_tmp)
    elif list_ActivityScoreMaster[idx_act]['Activity Type'] == 'Click Link':
        ii2                         = random.randint(0, len(list_ClickLinkSample) - 1)
        dict_tmp['LinkID']          = list_ClickLinkSample[ii2]['LinkID']
        dict_tmp['LinkIDValue']     = list_ClickLinkSample[ii2]['LinkIDValue']
        dict_tmp['QueryParameters'] = list_ClickLinkSample[ii2]['QueryParameters']
        dict_tmp['ReferrerURL']     = list_ClickLinkSample[ii2]['ReferrerURL']
        list_ClickLink.append(dict_tmp)
    elif list_ActivityScoreMaster[idx_act]['Activity Type'] == 'Open Email':
        ii2                         = random.randint(0, len(list_OpenEmailSample) - 1)
        dict_tmp['MailingID']       = list_OpenEmailSample[ii2]['MailingID']
        dict_tmp['MailingIDValue']  = list_OpenEmailSample[ii2]['MailingIDValue']
        list_OpenEmail.append(dict_tmp)
    elif list_ActivityScoreMaster[idx_act]['Activity Type'] == 'Click Email':
        ii2                         = random.randint(0, len(list_ClickEmailSample) - 1)
        dict_tmp['MailingID']       = list_ClickEmailSample[ii2]['MailingID']
        dict_tmp['MailingIDValue']  = list_ClickEmailSample[ii2]['MailingIDValue']
        dict_tmp['Link']            = list_ClickEmailSample[ii2]['Link']
        list_ClickEmail.append(dict_tmp)
    elif list_ActivityScoreMaster[idx_act]['Activity Type'] == 'Fill Out Form':
        ii2                         = random.randint(0, len(list_FillOutFormSample) - 1)
        dict_tmp['ProgramName']     = list_FillOutFormSample[ii2]['ProgramName']
        dict_tmp['ProgramType']     = list_FillOutFormSample[ii2]['ProgramType']
        dict_tmp['WebformID']       = list_FillOutFormSample[ii2]['WebformID']
        dict_tmp['WebformIDValue']  = list_FillOutFormSample[ii2]['WebformIDValue']
        list_FillOutForm.append(dict_tmp)


#   新規リードデータの書き出し
if list_LeadDataNew:
    pidbcommon.write_csv_overwrite(dict_param['file_New'], list_LeadDataNew)
#   リードデータへの追加
    if os.path.isfile(dict_configData['Leads_New']['file_Whole']):
        if os.path.getsize(dict_configData['Leads_New']['file_Whole']) > 0:
            pidbcommon.write_csv_appendwrite(dict_configData['Leads_New']['file_Whole'], list_LeadDataNew)
        else:
            pidbcommon.write_csv_overwrite(dict_configData['Leads_New']['file_Whole'], list_LeadDataNew)
    else:
        pidbcommon.write_csv_overwrite(dict_configData['Leads_New']['file_Whole'], list_LeadDataNew)
else:
    pathlib.Path(dict_param['file_New']).touch()

#   ChangeDataValueの書き出し
pidbcommon.write_csv(dict_configData['common']['actfileprefix'] + 'ChangeDataValue.csv', \
            dict_configData['common']['actfileprefix'] + 'ChangeDataValue_New.csv', \
            list_ChangeDataValue)

#   ClickEmail の書き出し
pidbcommon.write_csv(dict_configData['common']['actfileprefix'] + 'ClickEmail.csv', \
            dict_configData['common']['actfileprefix'] + 'ClickEmail_New.csv', \
            list_ClickEmail)

#   ClickLink の書き出し
pidbcommon.write_csv(dict_configData['common']['actfileprefix'] + 'ClickLink.csv', \
            dict_configData['common']['actfileprefix'] + 'ClickLink_New.csv', \
            list_ClickLink)

#   FillOutForm の書き出し
pidbcommon.write_csv(dict_configData['common']['actfileprefix'] + 'FillOutForm.csv', \
            dict_configData['common']['actfileprefix'] + 'FillOutForm_New.csv', \
            list_FillOutForm)

#   OpenEmail の書き出し
pidbcommon.write_csv(dict_configData['common']['actfileprefix'] + 'OpenEmail.csv', \
            dict_configData['common']['actfileprefix'] + 'OpenEmail_New.csv', \
            list_OpenEmail)

#   VisitWebpage の書き出し
pidbcommon.write_csv(dict_configData['common']['actfileprefix'] + 'VisitWebpage.csv', \
            dict_configData['common']['actfileprefix'] + 'VisitWebpage_New.csv', \
            list_VisitWebpage)



#   保存情報の登録
#if dt_baseDateCreated > dt_baseDateUpdated:
#    dict_configData['UpdatedAt']          = dt_baseDateCreated.isoformat()
dict_configData['Leads_New']['counter'] += len(list_LeadDataNew)
dt_procDate                             = dt.datetime.now()
dt_procDateJST                          = dt_procDate.replace(tzinfo=zoneinfo.ZoneInfo(key='Asia/Tokyo'))
dict_configData['Leads_New']['procdate']    = dt_procDateJST.isoformat().replace(r'\.\d{6)}', '')
with open(dict_param['configuration'], 'w') as f:
    json.dump(dict_configData, f, indent=4, ensure_ascii=False)

