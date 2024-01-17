import sys
import datetime as dt
import re
import csv
import random
import numpy as np
import json
from dateutil.relativedelta import relativedelta
import openpyxl
import pathlib
import os
import zoneinfo
import pprint
import pidbcommon


args                = sys.argv
flg_debug           = False
dict_param          = {}
baseDate            = ''
int_Number          = 30
dict_param['merge'] = True
dt_periodDate       = ''

for arg in args:
    print(arg)
    if arg == 'Leads_Renew.py':
        continue
    elif re.match(r'confi?g?u?r?e?=', arg, re.IGNORECASE):
        elements = re.split('=', arg)
        dict_param['configuration'] = elements[1]
    elif re.match(r'file_Renew=', arg, re.IGNORECASE):
        elements = re.split('=', arg)
        dict_param['file_Renew'] = elements[1]
    else:
        print("Unknown argument : "+arg, file=sys.stderr)
        sys.exit()

if 'configuration' in dict_param:
    #   保存情報の取得
    with open(dict_param['configuration']) as f:
        dict_configData = json.load(f)
    print('-------------------------------- dict_configData')
    pprint.pprint(dict_configData)
else:
    print("'conf(iguration)=' parameter not specified.", file=sys.stderr)
    sys.exit()

if 'file_Renew' in dict_param:
    pass
elif 'file_Renew' in dict_configData['Leads_Renew']:
    dict_param['file_Renew'] = dict_configData['Leads_Renew']['file_Renew']
else:
    print("'file_Renew' parameter is not found.", file=sys.stderr)
    sys.exit()

if not 'actfileprefix' in dict_configData['common']:
    print("'actfileprefix' key in JSON fike not found", file=sys.stderr)
    sys.exit()

if 'number' in dict_configData['Leads_Renew']:
    int_Number = dict_configData['Leads_Renew']['number']
else:
    print("'number' key in JSON file not found. 'number' was set to 30.")

if 'UpdatedAt' in dict_configData['Leads_Renew']:
    try:
        dt_tmp = dt.datetime.fromisoformat(dict_configData['Leads_Renew']['UpdatedAt'])
    except ValueError:
        print("Date format error : UpdatedAt: "+dict_configData['Leads_Renew']['UpdatedAt'], file=sys.stderr)
        sys.exit()
    str_UpdatedAt = dict_configData['Leads_Renew']['UpdatedAt'].replace('Z', '+00:00')
    jst_baseDate = dt.datetime.fromisoformat(str_UpdatedAt)
    td = dt.timedelta(days=1)
    jst_baseDate += td
    str_baseDay = jst_baseDate.date().isoformat()
    re.sub('T.*', '', baseDate)
else:
    print("'UpdatedAt' key in JSON fike not found", file=sys.stderr)
    sys.exit()


wb = openpyxl.load_workbook(dict_configData['common']['referenceSheet'])

#   行動スコアマスターを辞書のリストに格納
ws = wb['行動スコアマスター']
header_cells = ws[1]
list_ActivitiesScoreMaster = []
for row in ws.iter_rows(min_row=2, max_row=6):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_ActivitiesScoreMaster.append(row_dic)
print('-------------------------------- 行動スコアマスター')
print(list_ActivitiesScoreMaster)

ws = wb['ClickEmail']
header_cells = ws[1]
list_ClickEmailSample = []
for row in ws.iter_rows(min_row=2, max_row=59):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_ClickEmailSample.append(row_dic)

ws = wb['ClickLink']
header_cells = ws[1]
list_ClickLinkSample = []
for row in ws.iter_rows(min_row=2, max_row=355):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_ClickLinkSample.append(row_dic)

ws = wb['FillOutForm']
header_cells = ws[1]
list_FillOutFormSample = []
for row in ws.iter_rows(min_row=2, max_row=3717):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_FillOutFormSample.append(row_dic)

ws = wb['OpenEmail']
header_cells = ws[1]
list_OpenEmailSample = []
for row in ws.iter_rows(min_row=2, max_row=82):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_OpenEmailSample.append(row_dic)

ws = wb['VisitWebPage']
header_cells = ws[1]
list_VisitWebpageSample = []
for row in ws.iter_rows(min_row=2, max_row=497):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_VisitWebpageSample.append(row_dic)

#   既存リードデータを作成日と処理日の間隔をもとに別のリストに振り分けて格納する
dict_leadData = {'1':[], '3':[], '6':[], '12':[], '99':[], '00':[]}
dt_dateUTC  = dt.datetime.utcnow()
print('------------------------------------------------- dt_dateUTC')
print(dt_dateUTC)
dt_1ma      = jst_baseDate + relativedelta(months=-1)
dt_3ma      = jst_baseDate + relativedelta(months=-3)
dt_6ma      = jst_baseDate + relativedelta(months=-6)
dt_12ma     = jst_baseDate + relativedelta(months=-12)

print('------------------------------------------------- dates')
print(dt_1ma, dt_3ma, dt_6ma, dt_12ma)

#   リードデータの読み込み（リード作成日からの期間ごとに辞書に振り分ける）
if 'file_Whole' in dict_configData['Leads_New']:
    with open(dict_configData['Leads_New']['file_Whole'], encoding='utf-8', newline='') as f:
        csvreader = csv.DictReader(f)
        for row in csvreader:
            tmp_date = row['CreatedAt'].replace('Z', '+00:00')
            dt_CreatedAt = dt.datetime.fromisoformat(tmp_date)
            if dt_CreatedAt > jst_baseDate:
                dict_leadData['00'].append(row)
            else:
                if dt_CreatedAt >= dt_1ma:
                    dict_leadData['1'].append(row)
                elif dt_CreatedAt >= dt_3ma:
                    dict_leadData['3'].append(row)
                elif dt_CreatedAt >= dt_6ma:
                    dict_leadData['6'].append(row)
                elif dt_CreatedAt >= dt_12ma:
                    dict_leadData['12'].append(row)
                else:
                    dict_leadData['99'].append(row)
else:
    print("saveData_Lead error : 'Leads' key in JSON fike not found", file=sys.stderr)
    sys.exit()


#        list_LeadData = [row for row in csvreader]

#print('------------------------------------------------------- dict_leadData')
#print(dict_leadData)

dict_LifeCycle      = {10: 'MEL', 25:'MQL', 40:'SAL', 50:'SQL'}
list_saveId         = []
list_modyfiedLead   = []
repeatCount = random.randint(0, int_Number)
print(">>>> repeatCount : {:02}".format(repeatCount))
if dict_configData['Leads_Renew']['period'] == 1:
    list_dd = [0] * repeatCount
else:
    list_dd = [int(n/100) for n in pidbcommon.rand_ints_nodup(0, dict_configData['Leads_Renew']['period'] * 100, repeatCount)]
list_dd.sort()
list_ChangeDataValue    = []
list_VisitWebpage       = []
list_FillOutForm        = []
list_ClickLink          = []
list_OpenEmail          = []
list_ClickEmail         = []
list_SQL                = []
ii = 0
td = dt.timedelta(days=dict_configData['Leads_Renew']['period'])
jst_periodDate = jst_baseDate - td
while ii < repeatCount:
    print(">>>> ii : {:04}".format(ii))
    agoX = np.random.choice(['1', '3', '6', '12', '99'], p=[49/100, 30/100, 15/100, 5/100, 1/100])
    list_targetLead = dict_leadData[agoX]
    if not list_targetLead:
        continue
    idx = 0
    for jj in range(int_Number):
        idx = random.randint(0, len(list_targetLead) - 1)
        if list_targetLead[idx]['customer_id'] in list_saveId:
            continue
        dt_temp = dt.datetime.fromisoformat(list_targetLead[idx]['CreatedAt'])
        if dt_temp > jst_periodDate:
            continue
        list_saveId.append(list_targetLead[idx])
        break
    print('-------------------------------- list_targetLead')
    print(list_targetLead[idx])

    if list_targetLead[idx]['lifecycle_status'] == 'SQL':
        continue
    
    #   行動スコアマスターよりセット
    list_targetLead[idx]['BehaviorScore'] = random.choice(dict_configData['Leads_Renew']['score'])
    print(">>>> {:02}".format(list_targetLead[idx]['BehaviorScore']))
    if list_targetLead[idx]['BehaviorScore'] + int(list_targetLead[idx]['DemographicScore']) > 100:
        list_targetLead[idx]['BehaviorScore'] = 100 - int(list_targetLead[idx]['DemographicScore'])
    elif list_targetLead[idx]['BehaviorScore'] + int(list_targetLead[idx]['DemographicScore']) < 0:
        list_targetLead[idx]['BehaviorScore'] = 0
    list_targetLead[idx]['score'] = int(list_targetLead[idx]['BehaviorScore']) + int(list_targetLead[idx]['DemographicScore'])
    tmp_old_CDV = list_targetLead[idx]['lifecycle_status']
    for k, v in dict_LifeCycle.items():
        if int(list_targetLead[idx]['score']) < k:
            break
        else:
            list_targetLead[idx]['lifecycle_status'] = v
    hr = '{:0=2}:'.format(random.randint(7, 22))
    mm = '{:0=2}:'.format(random.randint(0, 59))
    ss = '{:0=2}.'.format(random.randint(0, 59))
    ms = '{:0=6}'.format(random.randint(0, 999999))
    td = dt.timedelta(days=list_dd[ii])
    dt_CreatedAt                        = jst_baseDate + td
    baseDateUpdated                     = dt_CreatedAt.date().isoformat()
    re.sub('T.*', '', baseDateUpdated)
    UpdatedAtJST                        = baseDateUpdated + 'T' + hr + mm + ss + ms
    print(UpdatedAtJST)
    dt_UpdatedAt                        = dt.datetime.fromisoformat(UpdatedAtJST)
    dt_UpdatedAtJST                     = dt_UpdatedAt.replace(tzinfo=zoneinfo.ZoneInfo(key='Asia/Tokyo'))
    dt_UpdatedAtUTC                     = dt_UpdatedAtJST.astimezone(dt.timezone.utc)

    list_targetLead[idx]['UpdatedAt']   = dt_UpdatedAtUTC.isoformat()
    dict_configData['Leads_Renew']['UpdatedAt'] = dt_UpdatedAtJST.isoformat()

    #   更新リードをリストにため込む
    list_modyfiedLead.append(list_targetLead[idx])

    #   lifecycle_statusがSQLになった更新リードをリストにため込む
    if list_targetLead[idx]['lifecycle_status'] == 'SQL':
        list_SQL.append(list_targetLead[idx])

    #   Activityの作成
    dict_tmp        = {}
    dict_tmp_CDV    = {'ActivityId' :'', 'ActivityDate' :'', 'AttributeName' :1292, 'AttributeNameValue' :'Lifecycle Status', 'CampaignId' :5218, 'LeadId' :'', 'NewValue' :'', 'OldValue' :'', 'Reason' :''}
    idx_act = random.randint(0, len(list_ActivitiesScoreMaster) - 1)
    dict_configData['common']['ActivityId']        += 1
    dict_tmp['ActivityId']          = dict_configData['common']['ActivityId']
    dict_tmp['ActivityDate']        = list_targetLead[idx]['UpdatedAt']
    dict_tmp['LeadId']              = list_targetLead[idx]['customer_id']
    if tmp_old_CDV != list_targetLead[idx]['lifecycle_status']:
        dict_tmp_CDV['ActivityId']          = dict_tmp['ActivityId']
        dict_tmp_CDV['ActivityDate']        = dict_tmp['ActivityDate']
        dict_tmp_CDV['AttributeNameValue']  = 'Lifecycle Status'
        dict_tmp_CDV['LeadId']              = list_targetLead[idx]['customer_id']
        dict_tmp_CDV['NewValue']            = list_targetLead[idx]['lifecycle_status']
        dict_tmp_CDV['OldValue']            = tmp_old_CDV
        dict_tmp_CDV['Reason']              = ''
        list_ChangeDataValue.append(dict_tmp_CDV)
    if list_ActivitiesScoreMaster[idx_act]['Activity Type'] == 'Visit Webpage':
        ii2                         = random.randint(0, len(list_VisitWebpageSample) - 1)
        dict_tmp['WebpageIDValue']  = list_VisitWebpageSample[ii2]['WebpageIDValue']
        dict_tmp['WebpageURL']      = list_VisitWebpageSample[ii2]['WebpageURL']
        list_VisitWebpage.append(dict_tmp)
        print('-------------------------------- Visit Webpage')
    elif list_ActivitiesScoreMaster[idx_act]['Activity Type'] == 'Click Link':
        ii2                         = random.randint(0, len(list_ClickLinkSample) - 1)
        dict_tmp['LinkID']          = list_ClickLinkSample[ii2]['LinkID']
        dict_tmp['LinkIDValue']     = list_ClickLinkSample[ii2]['LinkIDValue']
        dict_tmp['QueryParameters'] = list_ClickLinkSample[ii2]['QueryParameters']
        dict_tmp['ReferrerURL']     = list_ClickLinkSample[ii2]['ReferrerURL']
        list_ClickLink.append(dict_tmp)
        print('-------------------------------- Click Link')
    elif list_ActivitiesScoreMaster[idx_act]['Activity Type'] == 'Open Email':
        ii2                         = random.randint(0, len(list_OpenEmailSample) - 1)
        dict_tmp['MailingID']       = list_OpenEmailSample[ii2]['MailingID']
        dict_tmp['MailingIDValue']  = list_OpenEmailSample[ii2]['MailingIDValue']
        list_OpenEmail.append(dict_tmp)
        print('-------------------------------- Open Email')
    elif list_ActivitiesScoreMaster[idx_act]['Activity Type'] == 'Click Email':
        ii2                         = random.randint(0, len(list_ClickEmailSample) - 1)
        dict_tmp['MailingID']       = list_ClickEmailSample[ii2]['MailingID']
        dict_tmp['MailingIDValue']  = list_ClickEmailSample[ii2]['MailingIDValue']
        dict_tmp['Link']            = list_ClickEmailSample[ii2]['Link']
        list_ClickEmail.append(dict_tmp)
        print('-------------------------------- Click Email')
    elif list_ActivitiesScoreMaster[idx_act]['Activity Type'] == 'Fill Out Form':
        ii2                         = random.randint(0, len(list_FillOutFormSample) - 1)
        dict_tmp['ProgramName']     = list_FillOutFormSample[ii2]['ProgramName']
        dict_tmp['ProgramType']     = list_FillOutFormSample[ii2]['ProgramType']
        dict_tmp['WebformID']       = list_FillOutFormSample[ii2]['WebformID']
        dict_tmp['WebformIDValue']  = list_FillOutFormSample[ii2]['WebformIDValue']
        list_FillOutForm.append(dict_tmp)
        print('-------------------------------- Fill Out Form')
    print(dict_tmp)
    ii += 1

#   更新リードの書き出し
if list_modyfiedLead:
    with open(dict_param['file_Renew'], "w", newline="") as f1:
        fieldnames = [k for k in list_modyfiedLead[0]]
        dict_writer = csv.DictWriter(f1, fieldnames=fieldnames)
        dict_writer.writeheader()
        dict_writer.writerows(list_modyfiedLead)
else:
    f1 = pathlib.Path(dict_param['file_Renew'])
    f1.touch()

#   更新を含むリードデータを一次元の辞書リストに変換
list_LeadData = []
for v in dict_leadData.values():
    list_LeadData.extend(v)
#print('------------------------------------------------------ list_LeadData')
#print(list_LeadData)

#   更新を含むリードデータの書き出し
with open(dict_configData['Leads_New']['file_Whole'], "w", newline="") as f1:
    fieldnames = [k for k in list_LeadData[0]]
    dict_writer = csv.DictWriter(f1, fieldnames=fieldnames)
    dict_writer.writeheader()
    dict_writer.writerows(list_LeadData)

#   ChangeDataValueの書き出し
pidbcommon.write_csv(dict_configData['common']['actfileprefix'] + 'ChangeDataValue.csv', \
            dict_configData['common']['actfileprefix'] + 'ChangeDataValue_New.csv', \
            list_ChangeDataValue)

#   ClickEmail の書き出し
pidbcommon.write_csv(dict_configData['common']['actfileprefix'] + 'ClickEmail.csv', \
            dict_configData['common']['actfileprefix'] + 'ClickEmail_New.csv', \
            list_ClickEmail)

#   ClickLink の書き出し
pidbcommon.write_csv(dict_configData['common']['actfileprefix'] + 'ClickLink.csv', \
            dict_configData['common']['actfileprefix'] + 'ClickLink_New.csv', \
            list_ClickLink)

#   FillOutForm の書き出し
pidbcommon.write_csv(dict_configData['common']['actfileprefix'] + 'FillOutForm.csv', \
            dict_configData['common']['actfileprefix'] + 'FillOutForm_New.csv', \
            list_FillOutForm)

#   OpenEmail の書き出し
pidbcommon.write_csv(dict_configData['common']['actfileprefix'] + 'OpenEmail.csv', \
            dict_configData['common']['actfileprefix'] + 'OpenEmail_New.csv', \
            list_OpenEmail)

#   VisitWebpage の書き出し
pidbcommon.write_csv(dict_configData['common']['actfileprefix'] + 'VisitWebpage.csv', \
            dict_configData['common']['actfileprefix'] + 'VisitWebpage_New.csv', \
            list_VisitWebpage)

#   SQLになったリードを書き出す
pidbcommon.write_csv_add(dict_configData['Leads_Renew']['file_SQL'], list_SQL)

#   保存情報の登録
dict_configData['Leads_Renew']['counter'] += len(list_SQL)
dt_procDate                             = dt.datetime.now()
dt_procDateJST                          = dt_procDate.replace(tzinfo=zoneinfo.ZoneInfo(key='Asia/Tokyo'))
dict_configData['Leads_Renew']['procdate'] = dt_procDateJST.isoformat().replace(r'\.\d{6)}', '')
with open(dict_param['configuration'], 'w') as f:
    json.dump(dict_configData, f, indent=4, ensure_ascii=False)



