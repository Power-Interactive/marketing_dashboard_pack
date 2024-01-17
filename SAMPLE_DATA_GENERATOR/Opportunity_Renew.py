import os
import sys
import re
import json
import csv
import datetime as dt
import openpyxl
import random
import pprint
import zoneinfo
import numpy as np
import random as rd
import string
import pidbcommon

def rand_ints_nodup(a, b, k):
  ns = []
  while len(ns) < k:
    n = rd.randint(a, b)
    if not n in ns:
      ns.append(n)
  return ns

args                = sys.argv
flg_debug           = False
dict_param          = {}
dict_configData       = {}
int_Number          = 30
dict_param['merge'] = True

for arg in args:
    print(arg)
    if arg == 'Opportunity_Renew.py':
        continue
    elif re.match(r'confi?g?u?r?e?=', arg, re.IGNORECASE):
        elements = re.split('=', arg)
        dict_param['configuration'] = elements[1]
    elif re.match(r'file_Renew=', arg, re.IGNORECASE):
        elements = re.split('=', arg)
        dict_param['file_Renew'] = elements[1]
    else:
        print("Unknown argument : "+arg, file=sys.stderr)
        sys.exit()

if 'configuration' in dict_param:
    with open(dict_param['configuration']) as f:
        dict_configData = json.load(f)
    print('-------------------------------- dict_configData')
    pprint.pprint(dict_configData)
else:
    print("'conf(iguration)=' parameter not specified.", file=sys.stderr)
    sys.exit()

if not 'file_Whole' in dict_configData['Opportunity_New']:
    print("Configuration JSON file error : 'file_Whole' key in JSON file not found", file=sys.stderr)
    sys.exit()
if not 'file_history_New' in dict_configData['Opportunity_Renew']:
    print("Configuration JSON file error : 'file_history_New' key in JSON file not found", file=sys.stderr)
    sys.exit()
if not 'file_history_Whole' in dict_configData['Opportunity_Renew']:
    print("Configuration JSON file error : 'file_history_Whole' key in JSON file not found", file=sys.stderr)
    sys.exit()

if 'file_Renew' in dict_param:
    pass
else:
    if 'file_Renew' in dict_configData['Opportunity_Renew']:
        dict_param['file_Renew'] = dict_configData["Opportunity_Renew"]['file_Renew']
    else:
        print("'file_Renew' parameter is not found.", file=sys.stderr)
        sys.exit()


if 'number' in dict_configData['Opportunity_Renew']:
    int_Number = dict_configData['Opportunity_Renew']['number']
else:
    print("[information] Number of generate data : 'number' key in JSON file not found. 'number' was set to 30.")


if 'LastModifiedDate' in dict_configData['Opportunity_Renew']:
    try:
        dt_tmp = dt.datetime.fromisoformat(dict_configData['Opportunity_Renew']['LastModifiedDate'])
    except ValueError:
        print("Date format error : LastModifiedDate -> "+dict_configData['Opportunity_Renew']['LastModifiedDate'], file=sys.stderr)
        sys.exit()
    str_LastModifiedDate = dict_configData['Opportunity_Renew']['LastModifiedDate']
    dt_baseDate = dt.datetime.fromisoformat(str_LastModifiedDate)
    td = dt.timedelta(days=1)
    dt_baseDate += td
    str_baseDay = dt_baseDate.date().isoformat()
    re.sub('T.*', '', str_baseDay)
else:
    print("Configuration JSON file error : 'LastModifiedDate' key in JSON file not found", file=sys.stderr)
    sys.exit()

#   EXCELブックの読み込み
wb = openpyxl.load_workbook(dict_configData['common']['referenceSheet'])

#   アカウントマスターを辞書のリストに格納
ws = wb['sample_sfa.Account']
header_cells = ws[1]
list_AccountMaster = []
dict_AccountMaster = {}
for ii, row in enumerate(ws.iter_rows(min_row=2, max_row=61, min_col=1, max_col=14)):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
        list_AccountMaster.append(row_dic)
    if ii < 19:
        dict_AccountMaster[row_dic['ID']] = 'J1'
    elif ii < 40:
        dict_AccountMaster[row_dic['ID']] = 'J2'
    else:
        dict_AccountMaster[row_dic['ID']] = 'J3'
#print('---------------------------------------------------------------- dict_AccountMaster')
#pprint.pprint(dict_AccountMaster)

#   商談マスターの商談フェーズ（フォーキャスト）を辞書のリストに格納
ws = wb['商談フェーズ']
header_cells = ws[2]
list_OppForecast = []
for row in ws.iter_rows(min_row=3, max_row=9, min_col=1, max_col=3):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_OppForecast.append(row_dic)
print('---------------------------------------------------------------- 商談フェーズ（フォーキャスト）')
pprint.pprint(list_OppForecast)

#   商談マスターのお客様との接触状況（Contact Status）を辞書のリストに格納
ws = wb['お客様との接触状況']
header_cells = ws[2]
list_OppContactStatus   = []
for row in ws.iter_rows(min_row=3, max_row=8, min_col=1, max_col=3):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_OppContactStatus.append(row_dic)
list_StageName          = [row['ステータス'] for row in list_OppContactStatus]
print('---------------------------------------------------------------- お客様との接触状況')
pprint.pprint(list_OppContactStatus)
pprint.pprint(list_StageName)

#   商談データをアカウントのグレードごとに読み込む
list_OppData = []
dict_OppData = {'J1':[], 'J2':[], 'J3':[]}
if os.path.isfile(dict_configData['Opportunity_New']['file_Whole']):
    with open(dict_configData['Opportunity_New']['file_Whole'], encoding='utf-8', newline='') as f:
        csvreader = csv.DictReader(f)
        for row in csvreader:
            dict_OppData[dict_AccountMaster[row['account_id']]].append(row)
else:
    print("商談データが存在しないため処理をスキップ")
    sys.exit()
#print('---------------------------------------------------------------- dict_OppData')
#pprint.pprint(dict_OppData)

#   企業グレードごとに更新対象商談を更新頻度によりその件数を確定する
list_Acount = ['J1', 'J2', 'J3']
list_OpportunityHistory = []
list_OpportunityMod = []
repeatCount = random.randint(0, int_Number)
print('---------------------------------------------------------------- repeatCount')
pprint.pprint(repeatCount)
num_Sum = dict_configData['Opportunity_Renew']['classRate']['J1'] + \
            dict_configData['Opportunity_Renew']['classRate']['J2'] + \
            dict_configData['Opportunity_Renew']['classRate']['J3']
countJ1 = int(repeatCount * dict_configData['Opportunity_Renew']['classRate']['J1'] / num_Sum)
countJ2 = int(repeatCount * dict_configData['Opportunity_Renew']['classRate']['J2'] / num_Sum)
countJ3 = int(repeatCount * dict_configData['Opportunity_Renew']['classRate']['J3'] / num_Sum)
print('---------------------------------------------------------------- countJ1 countJ2 countJ3')
print(countJ1, countJ2, countJ3)
dict_Number = {'J1':countJ1, 'J2':countJ2, 'J3':countJ3}
print('---------------------------------------------------------------- dict_Number')
pprint.pprint(dict_Number) 
list_Hr = [int(n/1000) for n in rand_ints_nodup(0, 23000, countJ1 + countJ2 + countJ3)]
if dict_configData['Opportunity_Renew']['period'] == 1:
    list_dd = [0] * repeatCount
else:
    list_dd = [int(n/100) for n in pidbcommon.rand_ints_nodup(0, dict_configData['Opportunity_Renew']['period'] * 100, repeatCount)]
list_dd.sort()
print("------------------------------------------------------------------------- list_dd")
print(list_dd)
jj = 0
for Jx in list_Acount:
    #   更新対象の商談を抽出
    list_Index = list(range(len(dict_OppData[Jx])))
    print('---------------------------------------------------------------- list_Index')
    print(list_Index)
    list_Target = []
    if len(list_Index) < dict_Number[Jx]:
        list_Target = random.sample(list_Index, len(list_Index))
    else:
        list_Target = random.sample(list_Index, dict_Number[Jx])
    print('---------------------------------------------------------------- list_Target')
    print(list_Target)

    for ii in list_Target:
        print('---------------------------------------------------------------- StageName(old)')
        print(dict_OppData[Jx][ii]['StageName'])

        dt_createdDate  = dt.datetime.fromisoformat(dict_OppData[Jx][ii]['CreatedDate'])
        if dt_createdDate > dt_baseDate:
            continue
        if dict_OppData[Jx][ii]['StageName'] == 'Closed Won / Closed Lost':
            continue

        if dict_OppData[Jx][ii]['StageName'] == 'Not Contact':
#            dt_createdDate  = dt.datetime.fromisoformat(dict_OppData[Jx][ii]['CreatedDate'])
            dt_elapsedDay   = dt_baseDate - dt_createdDate
            print('---------------------------------------------------------------- dt_elapsedDay1')
            print(dt_elapsedDay)
            if dt_elapsedDay.days >= rd.choice(dict_configData['Opportunity_Renew']['nopDate']):
                dict_OppData[Jx][ii]['StageName']   = 'Contacted'
            else:
                continue
        elif dict_OppData[Jx][ii]['StageName'] == 'Contacted':
#            dt_createdDate      = dt.datetime.fromisoformat(dict_OppData[Jx][ii]['CreatedDate'])
            dt_elapsedDay       = dt_baseDate - dt_createdDate
            print('---------------------------------------------------------------- dt_elapsedDay2')
            print(dt_elapsedDay)
            if dt_elapsedDay.days >= rd.choice(dict_configData['Opportunity_Renew']['nopDate']):
                dict_OppData[Jx][ii]['StageName']   = 'Proposed'
            else:
                continue
        elif dict_OppData[Jx][ii]['StageName'] == 'Proposed':
            dt_lastModifiedDate = dt.datetime.fromisoformat(dict_OppData[Jx][ii]['LastModifiedDate'])
            dt_elapsedDay       = dt_baseDate - dt_lastModifiedDate
            print('---------------------------------------------------------------- dt_elapsedDay3')
            print(dt_elapsedDay)
            if dt_elapsedDay.days >= rd.choice(dict_configData['Opportunity_Renew']['nopDate']):
                dict_OppData[Jx][ii]['StageName']   = 'Quoted'
            else:
                continue
        elif dict_OppData[Jx][ii]['StageName'] == 'Quoted':
            dt_lastModifiedDate = dt.datetime.fromisoformat(dict_OppData[Jx][ii]['LastModifiedDate'])
            dt_elapsedDay       = dt_baseDate - dt_lastModifiedDate
            print('---------------------------------------------------------------- dt_elapsedDay4')
            print(dt_elapsedDay)
            if dt_elapsedDay.days >= rd.choice(dict_configData['Opportunity_Renew']['nopDate']):
                dict_OppData[Jx][ii]['StageName']   = 'Selected/Negociated'
            else:
                continue
        elif dict_OppData[Jx][ii]['StageName'] == 'Selected/Negociated':
            dt_lastModifiedDate = dt.datetime.fromisoformat(dict_OppData[Jx][ii]['LastModifiedDate'])
            dt_elapsedDay       = dt_baseDate - dt_lastModifiedDate
            print('---------------------------------------------------------------- dt_elapsedDay5')
            print(dt_elapsedDay)
            if dt_elapsedDay.days >= rd.choice(dict_configData['Opportunity_Renew']['nopDate']):
                dict_OppData[Jx][ii]['StageName']   = 'Closed Won / Closed Lost'
            else:
                continue
        print('---------------------------------------------------------------- StageName(new)')
        print(dict_OppData[Jx][ii]['StageName'])

        #   StageとProbabilityのセット
        old_ForecastCategory = dict_OppData[Jx][ii]['ForecastCategory']
        num_Probability = 0.0
        if dict_OppData[Jx][ii]['StageName'] == 'Not Contact' or \
           dict_OppData[Jx][ii]['StageName'] == 'Contacted':
            dict_OppData[Jx][ii]['ForecastCategory'] = 'Pipeline'
        elif dict_OppData[Jx][ii]['StageName'] == 'Proposed':
            dict_OppData[Jx][ii]['ForecastCategory'] = 'Upside'
        elif dict_OppData[Jx][ii]['StageName'] == 'Quoted':
            dict_OppData[Jx][ii]['ForecastCategory'] = rd.choice(['BestCase', 'Upside'])
        elif dict_OppData[Jx][ii]['StageName'] == 'Selected/Negociated':
            dict_OppData[Jx][ii]['ForecastCategory'] = rd.choice(['BestCase', 'Commit'])
        elif dict_OppData[Jx][ii]['StageName'] == 'Closed Won / Closed Lost':
            list_WL     = ['Won', 'Lost']
            list_Weigt  = [6, 1]
            dict_OppData[Jx][ii]['ForecastCategory'] = random.choices(list_WL, weights=list_Weigt)[0]
        
        #   タイムスタンプのセット
        hr  = '{:0=2}:'.format(random.randint(5, 23))
        mm  = '{:0=2}:'.format(rd.randint(0, 59))
        ss  = '{:0=2}.'.format(random.randint(0, 59))
        ms  = '{:0=6}'.format(random.randint(0, 999999))
        td = dt.timedelta(days=list_dd[jj])
        print("------------------------------------------------------------------ list_dd[jj]")
        print(list_dd[jj], jj)
        jj += 1
        print("-------------------------------------------------- jj")
        print(jj)
        dt_LastModifiedDate                   = dt_baseDate + td
        print("------------------------------------------------------------------ dt_LastModifiedDate")
        print(dt_LastModifiedDate)
        baseDate                            = dt_LastModifiedDate.date().isoformat()
        re.sub('T.*', '', baseDate)
        LastModifiedDateJST                   = baseDate + 'T' + hr + mm + ss + ms
        print("------------------------------------------------------------------ LastModifiedDateJST")
        print(LastModifiedDateJST)
        dt_LastModifiedDate       = dt.datetime.fromisoformat(LastModifiedDateJST)
        dt_LastModifiedDateJST    = dt_LastModifiedDate.replace(tzinfo=zoneinfo.ZoneInfo(key='Asia/Tokyo'))
        dt_LastModifiedDateUTC    = dt_LastModifiedDateJST.astimezone(dt.timezone.utc)
        dict_OppData[Jx][ii]['LastModifiedDate']                    = dt_LastModifiedDateUTC.isoformat()
        dict_configData['Opportunity_Renew']['LastModifiedDate']    = dt_LastModifiedDateJST.isoformat()
     
        list_OpportunityMod.append(dict_OppData[Jx][ii])

        #   OpportunityHistoryのセット
        dict_OpportunityHistory = {'Id':'', 'SystemModstamp':'', 'Amount':'', 'CloseDate':'', 'CreatedById':'', \
                                'CreatedDate':'', 'ExpectedRevenue':'', 'ForecastCategory':'', 'IsDeleted':'', \
                                'OpportunityId':'', 'Probability':'', 'StageName':''}
        if old_ForecastCategory == dict_OppData[Jx][ii]['ForecastCategory']:
            continue
        else:
            dict_configData['Opportunity_Renew']['HistoryId'] += 1
            dict_OpportunityHistory['Id']               = '{:0=18}{:<8}'.format(dict_configData['Opportunity_Renew']['HistoryId'], \
                                                ''.join(rd.choices(string.ascii_letters + string.digits, k=8)))
            dict_OpportunityHistory['SystemModstamp']   = dict_OppData[Jx][ii]['LastModifiedDate']
            dict_OpportunityHistory['Amount']           = dict_OppData[Jx][ii]['amount']
            dict_OpportunityHistory['CloseDate']        = dict_OppData[Jx][ii]['CloseDate']
            dict_OpportunityHistory['CreatedDate']      = dict_OppData[Jx][ii]['CreatedDate']
            dict_OpportunityHistory['ExpectedRevenue']  = dict_OppData[Jx][ii]['amount']
            dict_OpportunityHistory['ForecastCategory'] = dict_OppData[Jx][ii]['ForecastCategory']
            dict_OpportunityHistory['IsDeleted']        = 'false'
            dict_OpportunityHistory['OpportunityId']    = dict_OppData[Jx][ii]['opportunity_id']
            list_Probability = [x['確度'] for x in list_OppForecast if x['フェーズ'] == dict_OppData[Jx][ii]['ForecastCategory']]
            dict_OpportunityHistory['Probability']      = list_Probability[0] if len(list_Probability) else 0.0
            dict_OpportunityHistory['StageName']        = dict_OppData[Jx][ii]['StageName']

            print("---------------------------------------------------- dict_OpportunityHistory")
            print(dict_OpportunityHistory)
            list_OpportunityHistory.append(dict_OpportunityHistory)



#print('---------------------------------------------------------------- dict_OppData')
#pprint.pprint(dict_OppData)
#   商談データ更新分の書き出し
if list_OpportunityMod:
    with open(dict_param['file_Renew'], "w", newline="") as f1:
        fieldnames = [k for k in list_OpportunityMod[0]]
        dict_writer = csv.DictWriter(f1, fieldnames=fieldnames)
        dict_writer.writeheader()
        dict_writer.writerows(list_OpportunityMod)

#   商談データの書き出し
list_OppData = []
for v in dict_OppData.values():
    list_OppData.extend(v)
with open(dict_configData['Opportunity_New']['file_Whole'], "w", newline="") as f1:
    fieldnames = [k for k in list_OppData[0]]
    dict_writer = csv.DictWriter(f1, fieldnames=fieldnames)
    dict_writer.writeheader()
    dict_writer.writerows(list_OppData)

#   商談履歴データの書き出し
if list_OpportunityHistory:
    with open(dict_configData['Opportunity_Renew']['file_history_New'], "w", newline="") as f1:
        fieldnames = [k for k in list_OpportunityHistory[0]]
        dict_writer = csv.DictWriter(f1, fieldnames=fieldnames)
        dict_writer.writeheader()
        dict_writer.writerows(list_OpportunityHistory)
    pidbcommon.write_csv_add(dict_configData['Opportunity_Renew']['file_history_Whole'], list_OpportunityHistory)


#   保存情報の登録
dt_procDate                             = dt.datetime.now()
dt_procDateJST                          = dt_procDate.replace(tzinfo=zoneinfo.ZoneInfo(key='Asia/Tokyo'))
dict_configData['Opportunity_Renew']['procdate']  = dt_procDateJST.isoformat().replace(r'\.\d{6)}', '')
with open(dict_param['configuration'], 'w') as f:
    json.dump(dict_configData, f, indent=4, ensure_ascii=False)
