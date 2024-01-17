import os
import sys
import re
import json
import openpyxl
import random as rd
import numpy as np
import string
import datetime as dt
import zoneinfo
import csv
import pprint
import random
import pidbcommon
import string

args                = sys.argv
flg_debug           = False
dict_param          = {}
dict_configData       = {}
int_Number          = 30
dict_param['merge'] = True

for arg in args:
    print(arg)
    if arg == 'Opportunity_New.py':
        continue
    elif re.match(r'confi?g?u?r?e?=', arg, re.IGNORECASE):
        elements = re.split('=', arg)
        dict_param['configuration'] = elements[1]
    elif re.match(r'file_Whole=', arg, re.IGNORECASE):
        elements = re.split('=', arg)
        dict_param['file_Whole'] = elements[1]
    elif re.match(r'file_New=', arg, re.IGNORECASE):
        elements = re.split('=', arg)
        dict_param['file_New'] = elements[1]
    else:
        print("Unknown argument : "+arg, file=sys.stderr)
        sys.exit()

if 'configuration' in dict_param:
    with open(dict_param['configuration']) as f:
        dict_configData = json.load(f)
    print('-------------------------------- dict_configData')
    pprint.pprint(dict_configData)
else:
    print("'conf(iguration)=' parameter not specified.", file=sys.stderr)
    sys.exit()

if 'file_Whole' in dict_param:
    pass
elif 'file_Whole' in dict_configData['Opportunity_New']:
    dict_param['file_Whole'] = dict_configData["Opportunity_New"]['file_Whole']
else:
    print("'file_Whole' parameter is not found.", file=sys.stderr)
    sys.exit()

if 'file_New' in dict_param:
    pass
else:
    if 'file_New' in dict_configData['Opportunity_New']:
        dict_param['file_New'] = dict_configData["Opportunity_New"]['file_New']
    else:
        print("'file_New' parameter is not found.", file=sys.stderr)
        sys.exit()

print("------------------------------> {}".format(dict_param['file_New']))

if 'CreatedDate' in dict_configData['Opportunity_New']:
    try:
        dt_tmp = dt.datetime.fromisoformat(dict_configData['Opportunity_New']['CreatedDate'])
    except ValueError:
        print("Date format error : CreatedDate -> "+dict_configData['Opportunity_New']['CreatedDate'], file=sys.stderr)
        sys.exit()
    str_CreatedDate = dict_configData['Opportunity_New']['CreatedDate'].replace('Z', '+00:00')
    dt_baseDate = dt.datetime.fromisoformat(str_CreatedDate)
    td = dt.timedelta(days=1)
    dt_baseDate += td
    str_baseDay = dt_baseDate.date().isoformat()
    re.sub('T.*', '', str_baseDay)
    '''
    lastDate = dict_configData['CreatedDate'].replace('Z', '+00:00')
    print('-------------------------------- lastDate')
    print(lastDate)
    dt_lastDate = dt.datetime.fromisoformat(lastDate)
    td = dt.timedelta(days=1)
    dt_baseDate = dt_lastDate + td
    print('-------------------------------- last , next')
    print(dt_lastDate, dt_baseDate)
    baseDate = dt_baseDate.date().isoformat()
    re.sub('T.*', '', baseDate)
    '''
else:
    print("configuration error : 'CreatedDate' key not found", file=sys.stderr)
    sys.exit()

#   商談データの読み込み
list_OppData = []
if 'file_Whole' in dict_configData['Opportunity_New']:
    if os.path.isfile(dict_configData['Opportunity_New']['file_Whole']):
        with open(dict_configData['Opportunity_New']['file_Whole'], encoding='utf-8', newline='') as f:
            csvreader = csv.DictReader(f)
            list_OppData = [row for row in csvreader]
else:
    dict_param['merge'] = False
#print("---------------------------------------------------------------- dict_param['merge']")
#print(dict_param['merge'])

#   リードデータを読み込みsfa_contact_idを重複を除いてリストに格納
list_sfa_contact_id = []
if 'file_Whole' in dict_configData['Leads_New']:
    if os.path.isfile(dict_configData['Leads_New']['file_Whole']):
        with open(dict_configData['Leads_New']['file_Whole'], encoding='utf-8', newline='') as f:
            csvreader = csv.DictReader(f)
            list_sfa_contact_id = list(set([row['sfa_contact_id'] for row in csvreader]))
else:
    print("file_Whole file not found : {}".format(dict_configData['Leads_New']), file=sys.stderr)
    sys.exit()

#   SQLに到達したリードデータを読み込む
list_SQL = []
if 'file_SQL' in dict_configData['Leads_Renew']:
    if os.path.isfile(dict_configData['Leads_Renew']['file_SQL']):
        with open(dict_configData['Leads_Renew']['file_SQL'], encoding='utf-8', newline='') as f:
            csvreader = csv.DictReader(f)
            list_SQL = [row for row in csvreader]
        os.remove(dict_configData['Leads_Renew']['file_SQL'])
    else:
        print("SQLに到達したリードがないため処理をスキップ")
        sys.exit()
else:
    print("'file_SQL' file not specified in configuration JSON => ['Leads_Renew']['file_SQL']", file=sys.stderr)
    sys.exit()
#print("---------------------------------------------------------------- list_sfa_contact_id")
#print(list_sfa_contact_id)

wb = openpyxl.load_workbook(dict_configData['common']['referenceSheet'])

#   sample_sfa.Accountを辞書のリストに格納
ws = wb['sample_sfa.Account']
header_cells = ws[1]
list_AccountMaster = []
dict_AccountMaster = {'J1':[], 'J2':[], 'J3':[]}
for ii, row in enumerate(ws.iter_rows(min_row=2, max_row=61)):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
        list_AccountMaster.append(row_dic)
    if ii < 19:
        dict_AccountMaster['J1'].append(row_dic)
    elif ii < 40:
        dict_AccountMaster['J2'].append(row_dic)
    else:
        dict_AccountMaster['J3'].append(row_dic)

#   商談ソースマスターを辞書のリストに格納
ws = wb['商談ソースマスター']
header_cells = ws[1]
list_OppSourceMaster = []
for row in ws.iter_rows(min_row=2, max_row=47, min_col=1, max_col=3):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_OppSourceMaster.append(row_dic)

#   ユーザーマスターを辞書のリストに格納
ws = wb['ユーザーマスター']
header_cells = ws[1]
list_OppUser = []
for row in ws.iter_rows(min_row=2, max_row=17, min_col=1, max_col=10):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_OppUser.append(row_dic)

#   商談マスターの商談フェーズ（フォーキャスト）を辞書のリストに格納
ws = wb['商談フェーズ']
header_cells = ws[2]
list_OppForecast = []
for row in ws.iter_rows(min_row=3, max_row=9, min_col=1, max_col=3):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_OppForecast.append(row_dic)
print('---------------------------------------------------------------- 商談フェーズ（フォーキャスト）')
print(list_OppForecast)

#   商談マスターのお客様との接触状況（Contact Status）を辞書のリストに格納
ws = wb['お客様との接触状況']
header_cells = ws[2]
list_OppContactStatus = []
for row in ws.iter_rows(min_row=3, max_row=8, min_col=1, max_col=3):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_OppContactStatus.append(row_dic)
print('---------------------------------------------------------------- お客様との接触状況')
print(list_OppContactStatus)

#   商談サンプルデータを辞書のリストに格納
ws = wb['Opportunity']
header_cells = ws[1]
list_OppSample = []
for row in ws.iter_rows(min_row=2, max_row=468, min_col=1, max_col=21):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_OppSample.append(row_dic)
#print('---------------------------------------------------------------- 商談サンプルデータ')
#print(list_OppSample)

#   コンタクトサンプルデータを辞書のリストに格納
ws = wb['Contact']
header_cells = ws[1]
list_ContactSample = []
for row in ws.iter_rows(min_row=2, max_row=2104, min_col=1, max_col=16):
    row_dic = {}
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    list_ContactSample.append(row_dic)
#print('---------------------------------------------------------------- コンタクトサンプルデータ')
#print(list_ContactSample)

list_StaySQL    = []
list_Type       = ['新規', '既存']
list_StageName  = ['Contacted', 'Not Contact']
list_Weight     = [2, 8]
repeatCount     = len(list_SQL)
print('----------------------------------------------------------------------- repeatCount')
print(repeatCount)
int_oppRate     = random.choice(dict_configData['Opportunity_New']['genRate'])
int_Choice      = int(repeatCount * int_oppRate / 10)
list_TargetIdx  = random.sample(list(range(repeatCount)), int_Choice)
list_TargetIdx.sort()

print('----------------------------------------------------------------------- list_TargetIdx')
print(list_TargetIdx)
if dict_configData['Opportunity_New']['period'] == 1:
    list_dd = [0] * repeatCount
else:
    list_dd = [int(n/100) for n in pidbcommon.rand_ints_nodup(0, dict_configData['Opportunity_New']['period'] * 100, repeatCount)]
list_dd.sort()
print('----------------------------------------------------------------------- list_dd')
print(list_dd)
list_Forecast_Progress  = list(string.ascii_lowercase)
list_Opportunity = []
ii = 0
idx_Target = 0
for idx_Target in list_TargetIdx:
#   ターゲットまで読み飛ばす
    while ii < idx_Target:
        list_StaySQL.append(list_SQL[ii])
        ii += 1

#   商談化
    dict_OppData = {'opportunity_id' :'', 'LastModifiedDate' :'', 'account_id' :'', 'amount' :'', \
                    'CloseDate' :'', 'CreatedDate' :'', 'FiscalYear' :'', 'FiscalQuarter' :'', \
                    'LeadSource' :'', 'SourceDetail1' :'', 'SourceDetail2' :'', 'SourceDetail3' :'', \
                    'Name' :'', 'OwnerId' :'', 'StageName' :'', 'Type' :'', 'ForecastCategory' :'', \
                    'pi_opp_category' :'', 'loss_reason' :'', 'loss_reason_txt' :'', 'ContactId': ''}

    Jx  = np.random.choice(['J1', 'J2', 'J3'], p=[30/60, 20/60, 10/60])

    #   インクリメント、ランダム生成などによるセット
    dict_configData['Opportunity_New']['opportunity_id']  += 1
    tmp_int                         = dict_configData['Opportunity_New']['opportunity_id']
    tmp_opportunity_id              = f'{tmp_int:011}' + ''.join(rd.choices(string.ascii_letters + string.digits, k=7))
    dict_OppData['opportunity_id']  = tmp_opportunity_id[:14] + random.choice(list_Forecast_Progress) + tmp_opportunity_id[15:]
    hr = '{:0=2}:'.format(random.randint(7, 22))
    mm = '{:0=2}:'.format(random.randint(0, 59))
    ss = '{:0=2}.'.format(random.randint(0, 59))
    ms = '{:0=6}'.format(random.randint(0, 999999))
    td = dt.timedelta(days=list_dd[ii])
    dt_CreatedDate              = dt_baseDate + td
    str_CreatedDateJST          = dt_CreatedDate.date().isoformat() + 'T' + hr + mm + ss + ms
    print(">>> {}".format(str_CreatedDateJST))
    dt_CreatedDate              = dt.datetime.fromisoformat(str_CreatedDateJST)
    dt_CreatedDateJST           = dt_CreatedDate.replace(tzinfo=zoneinfo.ZoneInfo(key='Asia/Tokyo'))
    dt_CreatedDateUTC           = dt_CreatedDateJST.astimezone(dt.timezone.utc)
    dict_OppData['CreatedDate'] = dt_CreatedDateUTC.isoformat()
    dict_configData['Opportunity_New']['CreatedDate']   = dt_CreatedDateJST.isoformat()
    dict_OppData['amount']      = rd.randrange(dict_configData['Opportunity_New']['bias']['amount'][Jx]['min'], \
                                               dict_configData['Opportunity_New']['bias']['amount'][Jx]['max']) * 1000

    dt_start        = dt_CreatedDateUTC + dt.timedelta(weeks=dict_configData['Opportunity_New']['bias']['CloseDate'][Jx]['min'])
    dt_end          = dt_CreatedDateUTC + dt.timedelta(weeks=dict_configData['Opportunity_New']['bias']['CloseDate'][Jx]['max'])
    dt_delta        = dt_end - dt_start
    int_delta       = (dt_delta.days * 24 * 60 * 60) + dt_delta.seconds
    random_second   = rd.randrange(int_delta)
    dt_targetDate   = dt_start + dt.timedelta(seconds=random_second)
    dict_OppData['CloseDate']       = dt_targetDate.isoformat()

    dict_OppData['FiscalYear']  = np.random.choice([i+1 for i in range(12)], \
                                                    p=[10/100, 1/100, 3/100, 40/100, 3/100, 5/100, 2/100, \
                                                       1/100, 2/100, 30/100, 2/100, 1/100])

    #   リードデータよりセットする
    dict_OppData['account_id']  = list_SQL[ii]['sfa_account_id']
    dict_OppData['ContactId']   = list_SQL[ii]['sfa_contact_id']

    #   商談ソースマスターよりセットする
    idx                             = rd.randint(0, len(list_OppSourceMaster) - 1)
    dict_OppData['SourceDetail1']   = list_OppSourceMaster[idx]['OpprtunitySource']
    dict_OppData['SourceDetail3']   = list_OppSourceMaster[idx]['Detailed_Opp_Source']

    #   商談サンプルデータよりセットする
    idx                     = rd.randint(0, len(list_OppSample) - 1)
    dict_OppData['Name']    = list_OppSample[idx]['Name']

    #   ユーザーマスターよりセットする
    idx                     = rd.randint(0, len(list_OppUser) - 1)
    dict_OppData['OwnerId'] = list_OppUser[idx]['Id']

    #   初期値をセットする
    dict_OppData['ForecastCategory']    = 'Pipeline'
    dict_OppData['Type']                = random.choices(list_Type, k=1, weights=list_Weight)[0]
    dict_OppData['StageName']           = random.choices(list_StageName, k=1, weights=list_Weight)[0]

    list_Opportunity.append(dict_OppData)
    ii += 1
print('----------------------------------------------------------------------- idx_Target')
print(idx_Target)

while idx_Target < repeatCount:
    list_StaySQL.append(list_SQL[idx_Target])
    idx_Target += 1

pprint.pprint(list_StaySQL)


if list_Opportunity:
    #   新規商談データの書き出し
    with open(dict_param['file_New'], "w", newline="") as f1:
        fieldnames = [k for k in list_Opportunity[0]]
        dict_writer = csv.DictWriter(f1, fieldnames=fieldnames)
        dict_writer.writeheader()
        dict_writer.writerows(list_Opportunity)
    #   商談データへの追加
    pidbcommon.write_csv_add(dict_configData['Opportunity_New']['file_Whole'], list_Opportunity)


#   商談化に至らなかったSQL到達リードを書き戻す
if list_StaySQL:
    pidbcommon.write_csv_overwrite(dict_configData['Leads_Renew']['file_SQL'], list_StaySQL)

#   保存情報の登録
dict_configData['Opportunity_New']['counter'] += len(list_Opportunity)
dt_procDate                             = dt.datetime.now()
dt_procDateJST                          = dt_procDate.replace(tzinfo=zoneinfo.ZoneInfo(key='Asia/Tokyo'))
dict_configData['Opportunity_New']['procdate']  = dt_procDateJST.isoformat().replace(r'\.\d{6)}', '')
with open(dict_param['configuration'], 'w') as f:
    json.dump(dict_configData, f, indent=4, ensure_ascii=False)

